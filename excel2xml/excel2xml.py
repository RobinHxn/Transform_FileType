#!/usr/bin/python
# -*- coding:utf-8 -*-

"""
Python version: V3.7

Author: Robin&HXN

File version: V1.0

File name: excel2xml.py

Created on: 20181011

Resume: Transform buried_point Excel to Xml

"""

import datetime as dt
import openpyxl as opx
import xml.dom.minidom as xdm


def get_excel(file):
    """

    :param file: path of file
    :return: excel workbook
    """
    try:
        workbook = opx.load_workbook(file)
        return workbook
    except Exception as e:
        print(str(e))


def excel_to_xml(excel_path, xml_path):
    """

    :param excel_path: input path of excel
    :param xml_path: input path of xml
    :return: output xml
    """
    wb = get_excel(excel_path)
    ws = wb["event_list"]
    max_rows = ws.max_row
    max_columns = ws.max_column

    xml_file = xdm.Document()
    root_element = xml_file.createElement("BP_Project")
    xml_file.appendChild(root_element)

    for num_row in range(2, max_rows):
        l1_element = xml_file.createElement("Event")
        l1_element_key_p = u"%s" % "product_name"
        l1_element_value_p = ws.cell(row=num_row+1, column=1).value
        l1_element_key_pg = u"%s" % "page_name"
        l1_element_value_pg = ws.cell(row=num_row+1, column=2).value
        l1_element_key = u"%s" % "event_name"
        l1_element_value = ws.cell(row=num_row+1, column=3).value
        l1_element.setAttribute(l1_element_key, l1_element_value)
        l1_element.setAttribute(l1_element_key_pg, l1_element_value_pg)
        l1_element.setAttribute(l1_element_key_p, l1_element_value_p)
        for num_col in range(0, max_columns-3):
            event_element = xml_file.createElement(u"%s" % ws.cell(row=2, column=num_col+4).value)
            key1 = u"%s" % "name"
            value1 = ws.cell(row=1, column=num_col+4).value
            key2 = u"%s" % "value"
            value2 = ws.cell(row=num_row+1, column=num_col+4).value
            if isinstance(value1, int):
                value1 = str(value1)
            if isinstance(value2, int):
                value2 = str(value2)
            event_element.setAttribute(key1, value1)
            event_element.setAttribute(key2, value2)
            l1_element.appendChild(event_element)
        root_element.appendChild(l1_element)

    f = open(xml_path, "w")
    f.write(xml_file.toprettyxml())
    f.close()


if __name__ == '__main__':
    input_excel_name = input("Please input Excel Name:")
    excel_path = u"/Users/huangxingnai/to_excel/%s.xlsx" % input_excel_name
    xml_path = u"/Users/huangxingnai/to_xml/%s.xml" % input_excel_name
    excel_to_xml(excel_path, xml_path)
    print(dt.datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
